from django.contrib import admin
from .models import Members, Capacity, Selection, Shared,UserImportJob,CapacityImportJob
from django.contrib.auth.models import User
from django.utils.html import format_html
from django.urls import reverse
from django.utils.safestring import mark_safe
from django.db import transaction
from openpyxl import load_workbook
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

@admin.register(UserImportJob)
class UserImportJobAdmin(admin.ModelAdmin):
    list_display = ('id', 'status', 'result_preview')
    readonly_fields = ('status', 'result', 'import_file_template')
    
    def has_change_permission(self, request, obj=None):
        return False  # Disable editing records

    def has_delete_permission(self, request, obj=None):
        return False  # Disable deleting records
        
    list_filter = ('status',)
    fieldsets = (
        (None, {'fields': ('import_file_template', 'import_file')}),
        ('Processing Info', {
            'fields': ('status', 'result'),
            'classes': ('collapse',)
        }),
    )

    def result_preview(self, obj):
        return (obj.result[:75] + '...') if len(obj.result) > 75 else obj.result
    result_preview.short_description = 'Result'

    def import_file_template(self, obj):
        template_url = reverse('admin_download_template')
        instructions = """
        <div style="margin: 15px 0; padding: 10px; background: #f8f9fa; border-radius: 5px;">
            <h4 style="margin-top: 0;">Import Instructions</h4>
            <ol style="margin-bottom: 0;">
                <li>Download the template using the button below</li>
                <li>Keep the header row exactly as provided</li>
                <li>Fill in the data:
                    <ul>
                        <li><strong>Rank:</strong> Sequential number for line selection</li>
                        <li><strong>UserName:</strong> Must be student ID</li>
                        <li><strong>Password:</strong> Plain text password</li>
                        <li><strong>Name:</strong> Full name of student</li>
                        <li><strong>Start time/End time:</strong> Format as YYYY-MM-DD HH:MM:SS</li>
                    </ul>
                </li>
                <li>Save the file and upload it</li>
            </ol>
        </div>
        """
        download_btn = format_html(
            '<a href="{}" download class="button" style="margin-bottom: 15px;">'
            '<i class="fa fa-download"></i> Download Template</a>',
            template_url
        )
        return format_html("{}{}", mark_safe(instructions), download_btn)
    import_file_template.short_description = "Template & Instructions"

    def save_model(self, request, obj, form, change):
        if not change:  # Only process on initial creation
            obj.status = 'Processing'
            obj.result = 'Starting import process...'
            super().save_model(request, obj, form, change)
            
            # Use transaction.on_commit to schedule the processing after the current transaction
            transaction.on_commit(lambda: self.process_import_file(obj))
        else:
            super().save_model(request, obj, form, change)

    def process_import_file(self, job):
        try:
            with transaction.atomic():
                wb = load_workbook(job.import_file.file)
                ws = wb.active
                created_users = 0
                updated_users = 0
                created_members = 0
                updated_members = 0
                errors = []

                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):  # Skip empty rows
                        continue

                    try:
                        num, username, password, name, start_time, end_time = row
                        
                        # Validate required fields
                        if not all([num, username, password, name, start_time, end_time]):
                            errors.append(f"Row {row_num}: Missing required fields")
                            continue
                            
                        # Process datetime fields
                        try:
                            start_time = datetime.strptime(str(start_time), '%Y-%m-%d %H:%M:%S')
                            end_time = datetime.strptime(str(end_time), '%Y-%m-%d %H:%M:%S')
                        except ValueError as e:
                            errors.append(f"Row {row_num}: Invalid date format - {str(e)}")
                            continue

                        # Handle User
                        with transaction.atomic():
                            user, user_created = User.objects.get_or_create(
                                username=str(username).strip(),
                                defaults={
                                    'password': str(password)  # This will be hashed below
                                }
                            )
                            
                            if user_created:
                                user.set_password(str(password))
                                user.save()
                                created_users += 1
                            else:
                                # Update existing user's password
                                user.set_password(str(password))
                                user.save()
                                updated_users += 1
                            
                            # Handle Member
                            member_data = {
                                'num': num,
                                'name': str(name).strip(),
                                'start_time': start_time,
                                'end_time': end_time,
                                'logged': False,
                                'first': False,
                                'last': False
                            }
                            
                            member, member_created = Members.objects.update_or_create(
                                student_id=str(username).strip(),
                                defaults=member_data
                            )
                            
                            if member_created:
                                created_members += 1
                            else:
                                updated_members += 1

                    except Exception as e:
                        errors.append(f"Row {row_num}: Error - {str(e)}")
                        logger.error(f"Error processing row {row_num}: {str(e)}", exc_info=True)
                        continue

                # Build result message
                result = [
                    f"Processed {ws.max_row - 1} rows",
                    f"Created {created_users} new users",
                    f"Updated {updated_users} existing users",
                    f"Created {created_members} new member profiles",
                    f"Updated {updated_members} existing member profiles",
                    f"Errors encountered: {len(errors)}"
                ]
                
                if errors:
                    result.append("\nFirst 5 errors:")
                    result.extend(errors[:5])
                    if len(errors) > 5:
                        result.append(f"... ({len(errors) - 5} more errors)")

                job.status = "Completed"
                job.result = "\n".join(result)
                job.save()

        except Exception as e:
            logger.error(f"Import failed: {str(e)}", exc_info=True)
            job.status = "Failed"
            job.result = f"Critical error: {str(e)}"
            job.save(update_fields=['status', 'result'])


@admin.register(Members)
class MembersAdmin(admin.ModelAdmin):
    ordering = ('num',)
    list_display = ('id','num', 'name', 'student_id','start_time', 'end_time', 'logged', 'line')
    list_filter = ('logged', 'line')
    search_fields = ('name', 'student_id')
    list_editable = ('start_time', 'end_time','logged')
    readonly_fields = ('first', 'last')
    exclude = ('gender',)
    fieldsets = (
        ('Basic Information', {
            'fields': ('num', 'name', 'student_id')
        }),
        ('Status', {
            'fields': ('logged', 'first', 'last')
        }),
        ('Time Information', {
            'fields': ('start_time', 'end_time')
        }),
        ('Department Information', {
            'fields': ('line', 'surgery', 'internal', 'pediatrics', 
                      'gynocology', 'emergency', 'orthopedix', 'cardiology')
        }),
        ('Electives', {
            'fields': ('elective_one', 'elective_two', 'elective_three')
        }),
    )

@admin.register(Capacity)
class CapacityAdmin(admin.ModelAdmin):
    list_display = ('line', 'unit', 'hospital', 'full', 'reserved', 'remain')
    list_filter = ('line', 'unit', 'hospital')
    search_fields = ('unit', 'hospital')
    exclude = ('gender', 'common')
    list_editable = ('full', 'reserved', 'remain')


@admin.register(Shared)
class SharedAdmin(admin.ModelAdmin):
    list_display = ('unit', 'shared_unit')
    list_filter = ('unit',)
    search_fields = ('unit', 'shared_unit')